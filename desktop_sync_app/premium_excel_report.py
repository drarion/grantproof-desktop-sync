from __future__ import annotations

import tempfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment
from PIL import Image

from premium_report import PremiumActivityReportBuilder, resource_path, _safe_text

IMAGE_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.bmp', '.gif', '.webp'}


class PremiumExcelActivityReportBuilder:
    """Template-based premium XLSX activity report.

    The workbook layout is owned by the Excel template. This builder only replaces
    placeholders and inserts dynamic pictures/icons while preserving the template
    formatting as much as possible.
    """

    def __init__(self, base_folder: Path, default_org_name: str = '') -> None:
        self.base_folder = Path(base_folder)
        self.default_org_name = _safe_text(default_org_name)
        self.template_path = resource_path('desktop_sync_app/assets/templates/grantproof_report_template_excel.xlsx')
        self.docx_helper = PremiumActivityReportBuilder(base_folder, default_org_name=default_org_name)

    def build(self, output: Path, project_code: str, project_name: str, records: list[Any], lang: str) -> None:
        lang = 'en' if str(lang).lower().startswith('en') else 'fr'
        if not records:
            return
        output.parent.mkdir(parents=True, exist_ok=True)
        if not self.template_path.exists():
            raise FileNotFoundError(f'Excel report template not found: {self.template_path}')

        data = self.docx_helper._prepare_data(project_code, project_name, records, lang)
        with tempfile.TemporaryDirectory(prefix='grantproof_excel_report_assets_') as tmp:
            tmp_dir = Path(tmp)
            self.docx_helper._temp_dir = tmp_dir
            self.docx_helper._icon_cache = {}
            data['hero_picture'] = self.docx_helper._prepare_photo(data.get('hero_image'), tmp_dir / 'hero.png', (1100, 760))
            data['map_picture'] = self.docx_helper._render_map_image(data, tmp_dir / 'map.png', (980, 420), lang)
            data['header_country_picture'] = self.docx_helper._render_country_silhouette(data, tmp_dir / 'country.png', (260, 170), for_header=True)

            wb = load_workbook(self.template_path)
            self._fill_workbook(wb, data, lang, tmp_dir)
            self._finalize_workbook(wb)
            wb.save(output)
            self.docx_helper._temp_dir = None
            self.docx_helper._icon_cache = {}

    def _fill_workbook(self, wb, data: dict[str, Any], lang: str, tmp_dir: Path) -> None:
        kpis = self._normalized_kpis(data, lang)
        sectors = self._normalized_sectors(data)
        gps_text = self._gps_text(data, lang)
        mapping = {
            '{{REPORT_TITLE}}': data['title'].upper(),
            '{{PROJECT_NAME}}': data['project_name'],
            '{{PROJECT_CODE}}': data['project_code'],
            '{{ORGANIZATION}}': data['org_name'],
            '{{LOCATION}}': data['location'],
            '{{REPORT_DATE}}': data['report_date'],
            '{{KPI_1_VALUE}}': kpis[0].get('value', ''),
            '{{KPI_1_LABEL}}': kpis[0].get('label', ''),
            '{{KPI_2_VALUE}}': kpis[1].get('value', ''),
            '{{KPI_2_LABEL}}': kpis[1].get('label', ''),
            '{{KPI_3_VALUE}}': kpis[2].get('value', ''),
            '{{KPI_3_LABEL}}': kpis[2].get('label', ''),
            '{{SUMMARY}}': data['summary'],
            '{{TARGET_GROUP}}': data['target_group'],
            '{{EVIDENCE_TYPE}}': data['evidence_type'],
            '{{SECTOR_1_LABEL}}': sectors[0].get('label', ''),
            '{{SECTOR_2_LABEL}}': sectors[1].get('label', ''),
            '{{PROJECT_OUTPUT}}': data.get('activity_output') or ('Output projet à renseigner' if lang == 'fr' else 'Project output to be completed'),
            '{{MAP_LOCATION_LABEL}}': data.get('map_location_label') or data['location'],
            '{{GPS_COORDINATES}}': gps_text,
            '{{ANNEX_CAPTION_1}}': self._annex_caption(data, 0, lang),
            '{{ANNEX_CAPTION_2}}': self._annex_caption(data, 1, lang),
            '{{ANNEX_CAPTION_3}}': self._annex_caption(data, 2, lang),
            '{{ANNEX_CAPTION_4}}': self._annex_caption(data, 3, lang),
        }
        image_placeholders = {'{{MAP_IMAGE}}', '{{HERO_IMAGE}}', '{{HEADER_COUNTRY_MAP}}', '{{KPI_1_ICON}}', '{{KPI_2_ICON}}', '{{KPI_3_ICON}}', '{{SECTOR_1_ICON}}', '{{SECTOR_2_ICON}}', '{{ANNEX_IMAGE_1}}', '{{ANNEX_IMAGE_2}}', '{{ANNEX_IMAGE_3}}', '{{ANNEX_IMAGE_4}}'}
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        text = cell.value
                        if any(token in text for token in image_placeholders):
                            cell.value = ''
                            continue
                        for key, value in mapping.items():
                            if key in text:
                                text = text.replace(key, _safe_text(value))
                        cell.value = text
                        if '\n' in text or len(text) > 90:
                            cell.alignment = cell.alignment.copy(wrap_text=True)

        if 'Données placeholders' in wb.sheetnames:
            wb['Données placeholders'].sheet_state = 'hidden'
        if 'Rapport' in wb.sheetnames:
            self._fill_report_sheet(wb['Rapport'], data, kpis, sectors, tmp_dir)
        if 'Annexe documentaire' in wb.sheetnames:
            self._fill_annex_sheet(wb['Annexe documentaire'], data, tmp_dir, lang)

    def _fill_report_sheet(self, ws, data: dict[str, Any], kpis: list[dict[str, str]], sectors: list[dict[str, str]], tmp_dir: Path) -> None:
        self._remove_images_at(ws, {(0, 6), (30, 6), (6, 1), (6, 2), (6, 3), (3, 4), (21, 1), (21, 2)})
        self._add_image(ws, data['header_country_picture'], 'G1', 87, 79)
        self._add_image(ws, data['header_country_picture'], 'G31', 87, 79)
        self._add_image(ws, data['hero_picture'], 'E4', 319, 278)
        self._add_image(ws, data['map_picture'], 'F22', 315, 135)

        kpi_cells = ['B7', 'C7', 'D7']
        for idx, kpi in enumerate(kpis[:3]):
            icon_path = self.docx_helper._plain_icon(self.docx_helper._kpi_icon_key(kpi))
            self._add_image(ws, icon_path, kpi_cells[idx], 61, 61)

        sector_cells = ['B22', 'C22']
        for idx, sector in enumerate(sectors[:2]):
            icon_key = 'food_security' if sector.get('key') in {'food_security', 'food'} else ('agriculture' if sector.get('key') == 'agriculture' else sector.get('icon', 'coordination'))
            icon_path = self.docx_helper._sector_icon(icon_key)
            self._add_image(ws, icon_path, sector_cells[idx], 59, 59)

        self._fill_inline_annex_on_report(ws, data, tmp_dir)

    def _fill_inline_annex_on_report(self, ws, data: dict[str, Any], tmp_dir: Path) -> None:
        images = [Path(p) for p in data.get('annex_images') or [] if Path(p).exists()]
        targets = [('A35', 250, 155), ('E35', 250, 155), ('A41', 250, 155)]
        for idx, (cell, width, height) in enumerate(targets):
            if idx < len(images):
                prepared = self.docx_helper._prepare_photo(images[idx], tmp_dir / f'inline_annex_{idx + 1}.png', (900, 560))
                self._add_image(ws, prepared, cell, width, height)

    def _fill_annex_sheet(self, ws, data: dict[str, Any], tmp_dir: Path, lang: str) -> None:
        images = [Path(p) for p in data.get('annex_images') or [] if Path(p).exists()]
        image_targets = [('A4', 385, 250), ('F4', 385, 250), ('A18', 385, 250), ('F18', 385, 250)]
        caption_cells = ['A17', 'F17', 'A31', 'F31']
        for idx, target in enumerate(image_targets):
            if idx < len(images):
                prepared = self.docx_helper._prepare_photo(images[idx], tmp_dir / f'annex_{idx + 1}.png', (1000, 650))
                self._add_image(ws, prepared, target[0], target[1], target[2])
            elif idx < len(caption_cells):
                ws[caption_cells[idx]] = ''

    def _normalized_kpis(self, data: dict[str, Any], lang: str) -> list[dict[str, str]]:
        kpis = list(data.get('kpis') or [])
        kinds = {item.get('kind') for item in kpis}
        if 'media' not in kinds:
            kpis.append({'kind': 'media', 'value': str(data.get('media_count', 0)), 'label': 'Médias liés' if lang == 'fr' else 'Linked media', 'icon': 'media'})
        if 'evidence' not in {item.get('kind') for item in kpis}:
            kpis.append({'kind': 'evidence', 'value': str(data.get('evidence_count', 1)), 'label': 'Preuve consolidée' if lang == 'fr' else 'Consolidated evidence', 'icon': 'check'})
        preferred = ['beneficiary', 'beneficiaries', 'people', 'farmer', 'kit', 'kits', 'group', 'groups', 'media', 'evidence']
        def rank(item: dict[str, str]) -> tuple[int, int]:
            kind = _safe_text(item.get('kind')).lower()
            return (preferred.index(kind) if kind in preferred else 99, -len(_safe_text(item.get('value'))))
        result: list[dict[str, str]] = []
        seen = set()
        for item in sorted(kpis, key=rank):
            sig = (item.get('kind'), item.get('value'))
            if sig in seen:
                continue
            result.append(item)
            seen.add(sig)
            if len(result) == 3:
                break
        while len(result) < 3:
            result.append({'kind': 'evidence', 'value': '—', 'label': 'Indicateur' if lang == 'fr' else 'Indicator', 'icon': 'check'})
        return result

    def _normalized_sectors(self, data: dict[str, Any]) -> list[dict[str, str]]:
        sectors = list(data.get('sectors') or [])
        sectors = sorted(sectors[:2], key=lambda s: 0 if s.get('key') in {'food_security', 'food'} else (1 if s.get('key') == 'agriculture' else 2))
        while len(sectors) < 2:
            sectors.append({'key': 'coordination', 'label': 'Coordination', 'icon': 'coordination'})
        return sectors

    def _gps_text(self, data: dict[str, Any], lang: str) -> str:
        if data.get('gps_point'):
            lat, lon, _ = data['gps_point']
            return f'GPS : {lat:.4f}, {lon:.4f}'
        return 'Carte nationale affichée par défaut.' if lang == 'fr' else 'National map shown by default.'

    def _annex_caption(self, data: dict[str, Any], index: int, lang: str) -> str:
        images = data.get('annex_images') or []
        if index >= len(images):
            return ''
        label = 'Média complémentaire' if lang == 'fr' else 'Additional media'
        return f'{label} {index + 1} — {data.get("activity") or data.get("location")}'

    def _remove_images_at(self, ws, anchors: set[tuple[int, int]]) -> None:
        kept = []
        for image in getattr(ws, '_images', []):
            try:
                pos = (image.anchor._from.row, image.anchor._from.col)
            except Exception:
                kept.append(image)
                continue
            if pos not in anchors:
                kept.append(image)
        ws._images = kept

    def _add_image(self, ws, path: Path | str | None, cell: str, width: int, height: int) -> None:
        if path is None or not Path(path).exists():
            return
        img = XLImage(str(path))
        img.width = int(width)
        img.height = int(height)
        ws.add_image(img, cell)

    def _finalize_workbook(self, wb) -> None:
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and ('{{' in cell.value or '}}' in cell.value):
                        cell.value = cell.value.replace('{{', '').replace('}}', '')
            ws.sheet_view.showGridLines = False
